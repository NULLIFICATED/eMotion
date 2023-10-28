from p5 import *
import openpyxl
import threading
mode = 1

def orientation(p, q, r):
    val = (q[1] - p[1]) * (r[0] - q[0]) - (q[0] - p[0]) * (r[1] - q[1])
    if val == 0:
        return 0
    if val > 0:
        return 1
    else:
        return 2
def is_cross(p1, q1, p2, q2):
    o1 = orientation(p1, q1, p2)
    o2 = orientation(p1, q1, q2)
    o3 = orientation(p2, q2, p1)
    o4 = orientation(p2, q2, q1)

    if o1 != o2 and o3 != o4:
        return True

    return False
def in_range(xpoint, ypoint, x1, y1, x2, y2):
    if (xpoint>x1)and(xpoint<x2)and(ypoint>y1)and(ypoint<y2):
        return True
    return False

class exponat():
    def __init__(self, x, y, w, h, name, desc):
        self.x = x
        self.y = y
        self.w = w
        self.h = h
        self.name = name
        self.desc = desc
        self.emotions = [0, 0, 0, 0, 0, 0]
        self.freq = 0
    def drow(self):
        stroke_weight(1)
        stroke(0)
        fill(100)
        rect((self.x-self.w/2, self.y-self.h/2), self.w, self.h)

class wall():
    def __init__(self, x1, y1, x2, y2):
        self.x1 = x1
        self.y1 = y1
        self.x2 = x2
        self.y2 = y2
    def drow(self):
        stroke_weight(10)
        stroke(0)
        line((self.x1, self.y1),(self.x2, self.y2))


exponats = []
walls = []
trace = []
traces = []
drawmode = 1
mode = 1


def save():
    name = 'n'
    wb = openpyxl.load_workbook('maps.xlsx')
    wb.create_sheet(name)
    ws = wb[name]
    for i in range(1, len(exponats)+1):
        ws.cell(row=i, column=0+1).value = exponats[i-1].name
        ws.cell(row=i, column=1+1).value = exponats[i-1].desc
        ws.cell(row=i, column=2+1).value = exponats[i-1].x
        ws.cell(row=i, column=3+1).value = exponats[i-1].y
        for j in range(6):
            ws.cell(row=i, column=4+1+j).value = exponats[i-1].emotions[j]
        ws.cell(row=i, column=10+1).value = exponats[i-1].freq
    for i in range(1, len(walls)+1):
        ws.cell(row=i, column=11+1).value = walls[i-1].x1
        ws.cell(row=i, column=12+1).value = walls[i-1].y1
        ws.cell(row=i, column=13+1).value = walls[i-1].x2
        ws.cell(row=i, column=14+1).value = walls[i-1].y2

    wb.save('maps.xlsx')

def load():
    name = 'n'
    wb = openpyxl.load_workbook('maps.xlsx')
    ws = wb[name]
    i=1
    while not(ws.cell(row=i, column=0+1).value)==None:
        exponats.append(exponat(ws.cell(row=i, column=2+1).value,ws.cell(row=i, column=3+1).value, 50,50, ws.cell(row=i, column=0+1).value, ws.cell(row=i, column=1+1).value))
        exponats[-1].emotions = [ws.cell(row=i, column=4+1).value, ws.cell(row=i, column=5+1).value, ws.cell(row=i, column=6+1).value, ws.cell(row=i, column=7+1).value, ws.cell(row=i, column=8+1).value, ws.cell(row=i, column=9+1).value]
        exponats[-1].freq = ws.cell(row=i, column=10+1).value
        i+=1
    i=1
    while not(ws.cell(row=i, column=11+1).value)==None:
        walls.append(wall(ws.cell(row=i, column=11+1).value, ws.cell(row=i, column=12+1).value, ws.cell(row=i, column=13+1).value, ws.cell(row=i, column=14+1).value))
    wb.save('maps.xlsx')





#UI begin
from PyQt5 import QtCore, QtGui, QtWidgets
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1100, 200)
        MainWindow.setMinimumSize(QtCore.QSize(1100, 200))
        MainWindow.setMaximumSize(QtCore.QSize(1100, 200))
        MainWindow.setStyleSheet("background-color: rgb(107, 148, 235);\n"
"background-color: rgb(0, 70, 132);")
        MainWindow.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setStyleSheet("color: rgb(0, 75, 137);")
        self.centralwidget.setObjectName("centralwidget")
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setEnabled(True)
        self.groupBox.setGeometry(QtCore.QRect(70, 0, 351, 131))
        self.groupBox.setStyleSheet("color: rgb(224, 227, 224);\n"
"font: 75 16pt \"Century Gothic\";")
        self.groupBox.setAlignment(QtCore.Qt.AlignCenter)
        self.groupBox.setFlat(False)
        self.groupBox.setCheckable(False)
        self.groupBox.setObjectName("groupBox")
        self.btn2 = QtWidgets.QRadioButton(self.groupBox)
        self.btn2.setGeometry(QtCore.QRect(10, 60, 331, 31))
        self.btn2.setStyleSheet("font: 14pt \"MS Shell Dlg 2\";")
        self.btn2.setObjectName("btn2")
        self.btn3 = QtWidgets.QRadioButton(self.groupBox)
        self.btn3.setGeometry(QtCore.QRect(10, 90, 261, 31))
        self.btn3.setStyleSheet("font: 14pt \"MS Shell Dlg 2\";")
        self.btn3.setObjectName("btn3")
        self.btn1 = QtWidgets.QRadioButton(self.groupBox)
        self.btn1.setGeometry(QtCore.QRect(10, 30, 311, 31))
        self.btn1.setStyleSheet("font: 14pt \"MS Shell Dlg 2\";")
        self.btn1.setObjectName("btn1")
        self.groupBox_2 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_2.setEnabled(False)
        self.groupBox_2.setGeometry(QtCore.QRect(420, 0, 321, 131))
        self.groupBox_2.setStyleSheet("color: rgb(224, 227, 224);\n"
"font: 75 16pt \"Century Gothic\";")
        self.groupBox_2.setObjectName("groupBox_2")
        self.btn4 = QtWidgets.QRadioButton(self.groupBox_2)
        self.btn4.setGeometry(QtCore.QRect(30, 40, 141, 20))
        self.btn4.setStyleSheet("font: 14pt \"MS Shell Dlg 2\";")
        self.btn4.setObjectName("btn4")
        self.btn5 = QtWidgets.QRadioButton(self.groupBox_2)
        self.btn5.setGeometry(QtCore.QRect(30, 70, 141, 20))
        self.btn5.setStyleSheet("font: 14pt \"MS Shell Dlg 2\";")
        self.btn5.setObjectName("btn5")
        self.groupBox_3 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_3.setGeometry(QtCore.QRect(740, 0, 341, 131))
        font = QtGui.QFont()
        font.setFamily("Century Gothic")
        font.setPointSize(16)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.groupBox_3.setFont(font)
        self.groupBox_3.setStyleSheet("font: 75 16pt \"Century Gothic\";\n"
"color: rgb(226, 229, 226);\n"
"")
        self.groupBox_3.setAlignment(QtCore.Qt.AlignCenter)
        self.groupBox_3.setObjectName("groupBox_3")
        self.savebtn = QtWidgets.QPushButton(self.groupBox_3)
        self.savebtn.setGeometry(QtCore.QRect(40, 40, 261, 41))
        font = QtGui.QFont()
        font.setFamily("Century Gothic")
        font.setPointSize(16)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.savebtn.setFont(font)
        self.savebtn.setObjectName("savebtn")
        self.loadbtn = QtWidgets.QPushButton(self.groupBox_3)
        self.loadbtn.setGeometry(QtCore.QRect(40, 80, 261, 41))
        font = QtGui.QFont()
        font.setFamily("Century Gothic")
        font.setPointSize(16)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.loadbtn.setFont(font)
        self.loadbtn.setObjectName("loadbtn")
        self.pbt1 = QtWidgets.QPushButton(self.centralwidget)
        self.pbt1.setEnabled(False)
        self.pbt1.setGeometry(QtCore.QRect(70, 140, 351, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pbt1.setFont(font)
        self.pbt1.setStyleSheet("color: rgb(255, 255, 255);")
        self.pbt1.setObjectName("pbt1")
        self.pbt2 = QtWidgets.QPushButton(self.centralwidget)
        self.pbt2.setGeometry(QtCore.QRect(490, 147, 171, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pbt2.setFont(font)
        self.pbt2.setStyleSheet("color: rgb(255, 255, 255);")
        self.pbt2.setObjectName("pbt2")
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        self.btn1.toggled['bool'].connect(self.groupBox_2.setEnabled) # type: ignore
        self.btn2.toggled['bool'].connect(self.pbt1.setEnabled) # type: ignore
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.btn1.toggled.connect(self.mode1)
        self.btn2.toggled.connect(self.mode2)
        self.btn3.toggled.connect(self.mode3)
        self.btn4.toggled.connect(self.drawmode1)
        self.btn5.toggled.connect(self.drawmode2)
        self.pbt1.clicked.connect(self.cleartraces)
        self.pbt2.clicked.connect(self.clearall)
        self.savebtn.clicked.connect(self.savestream)
        self.loadbtn.clicked.connect(self.loadstream)

    def savestream(self):
        thr1 = threading.Thread(target=save, name='t1')
        thr1.start()
    def loadstream(self):
        thr2 = threading.Thread(target=load, name='t2')
        thr2.start()
    def mode1(self):
        global mode
        mode = 1

    def mode2(self):
        global mode
        mode = 2

    def mode3(self):
        global mode
        mode = 3

    def drawmode1(self):
        global drawmode
        drawmode = 1

    def drawmode2(self):
        global drawmode
        drawmode = 2

    def cleartraces(self):
        global traces
        global trace
        traces.append(trace)
        for i in range(len(trace)):
            trace.pop()

    def clearall(self):
        global exponats
        global walls
        global trace
        exponats = []
        walls = []
        trace = []

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Панель инструментов"))
        self.groupBox.setTitle(_translate("MainWindow", "Режимы:"))
        self.btn2.setText(_translate("MainWindow", "Отслеживание траектории ➭"))
        self.btn3.setText(_translate("MainWindow", "Просмотр 👁️"))
        self.btn1.setText(_translate("MainWindow", "Создание карты ✎"))
        self.groupBox_2.setTitle(_translate("MainWindow", "Создаваемый объект:"))
        self.btn4.setText(_translate("MainWindow", "Экспонат"))
        self.btn5.setText(_translate("MainWindow", "Стена"))
        self.groupBox_3.setTitle(_translate("MainWindow", "Сохранение/загрузка"))
        self.savebtn.setText(_translate("MainWindow", "Сохранить"))
        self.loadbtn.setText(_translate("MainWindow", "Загрузить"))
        self.pbt1.setText(_translate("MainWindow", "Сохранить и очистить траектроию"))
        self.pbt2.setText(_translate("MainWindow", "Очистить всё"))

import sys
app = QtWidgets.QApplication(sys.argv)
MainWindow = QtWidgets.QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(MainWindow)
MainWindow.show()
#sys.exit(app.exec_())
#UI end





#Proxessing begin
def setup():
    size(1600,100)
def draw():
    background(200)
    for e in exponats:
        e.drow()
    for w in walls:
        w.drow()
    for i in range(len(trace)-1):
        stroke_weight(1)
        stroke(0)
        line(trace[i],trace[i+1])
    if mouse_is_pressed and mode == 2:
        if len(trace)==0:
            trace.append((mouse_x, mouse_y))
        else:
            if ((mouse_x-trace[-1][0])**2+(mouse_y-trace[-1][1])**2)>=30**2:
                trace.append((mouse_x, mouse_y))
    for i in range(len(walls)):
        if len(trace)>1:
            if is_cross([walls[i].x1, walls[i].y1],[walls[i].x2, walls[i].y2],trace[-2], trace[-1]):
                trace.pop()
def mouse_pressed():
    if mode == 1:
        if drawmode == 1:
            print('Введите название:')
            k = input()
            print('Введите описание:')
            l = input()
            exponats.append(exponat(mouse_x, mouse_y, 50, 50, k, l))
        if drawmode == 2:
            walls.append(wall(mouse_x, mouse_y, mouse_x, mouse_y))
    if mode == 3:
        for i in range(len(exponats)):
            if in_range(mouse_x,mouse_y,exponats[i].x,exponats[i].y,exponats[i].x+exponats[i].w,exponats[i].y+exponats[i].h):
                print("Название: ", exponats[i].name)
                print("Описание: ", exponats[i].desc)
                print('Восторг: ', exponats[i].emotions[0])
                print('Удовлетворение: ', exponats[i].emotions[1])
                print('Удивление: ', exponats[i].emotions[2])
                print('Нейтрален: ', exponats[i].emotions[3])
                print('Грусть: ', exponats[i].emotions[4])
                print('Злость: ', exponats[i].emotions[5])
                print('Количество посещений: ', exponats[i].freq)


def mouse_released():
    if mode == 1:
        if drawmode == 2:
            walls[-1].x2 = mouse_x
            walls[-1].y2 = mouse_y
    if mode == 2:
        for i in range(len(exponats)):
            if in_range(mouse_x,mouse_y,exponats[i].x,exponats[i].y,exponats[i].x+exponats[i].w,exponats[i].y+exponats[i].h):
                exponats[i].freq+=1
                print('Введите эмоцию из списка: \nВосторг\nУдовлетворение\nУдивление\nНейтрален\nГрусть\nЗлость')
                k = input()
                if k == 'Восторг':
                    exponats[i].emotions[0]+=1
                if k == 'Удовлетворение':
                    exponats[i].emotions[1]+=1
                if k == 'Удивление':
                    exponats[i].emotions[2]+=1
                if k == 'Нейтрален':
                    exponats[i].emotions[3]+=1
                if k == 'Грусть':
                    exponats[i].emotions[4]+=1
                if k == 'Злость':
                    exponats[i].emotions[5]+=1




#Processing end

run()
